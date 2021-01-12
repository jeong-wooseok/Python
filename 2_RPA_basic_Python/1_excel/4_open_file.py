from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook('sample.xlsx')
ws = wb.active

# cell 데이터 불러오기
# for x in range(1,11):
#     for y in range(1,11):
#         print(ws.cell(x,y).value, end=" ")
#     print()

# cell 갯수를 모를때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(x,y).value, end = " ")
    print()

#전체 rows
print(tuple(ws.rows))

for row in tuple(ws.rows):
    print(row[1].value) # 영어만 가져올 경우

#전체 columns
print(tuple(ws.columns)) 

for column in tuple(ws.columns):
    print(column[1].value) # 1번 학생의 점수

for row in ws.iter_rows(): #전체 row
    print(row[2].value)

for col in ws.iter_cols():
    print(col[0].value)

#범위지정 : 1~5줄까지, 2~3열까지
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3): 
    print(row[0].value,row[1].value)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=2, max_col=3):
    print(col)


