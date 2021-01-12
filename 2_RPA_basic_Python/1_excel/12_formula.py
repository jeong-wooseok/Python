import datetime
from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.create_sheet('fomula')

ws['A1'] = datetime.datetime.today()
ws['A2'] = '=SUM(1,2,3)'
ws['A3'] = '=AVERAGE(1,2,3)'
ws['A4'] = 10
ws['A5'] = 20
ws['A6'] = '=sum(A4:A5)'

wb.save('sample_modi.xlsx')