from openpyxl import load_workbook
wb = load_workbook('sample_modi.xlsx')
ws = wb.create_sheet('merge')

# merge
ws.merge_cells('B2:D2')
ws['B2'].value = 'merged cell'

wb.save('sample_modi.xlsx')