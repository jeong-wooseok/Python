from openpyxl import Workbook

wb=Workbook()
ws =wb.active
ws.title = 'MySheet'

# 각 셀에 값을 입력
for i in range(1:7):
    ws[f"A{i}"] = i


# ws['A1'] = 1
# ws['A2'] = 2
# ws['A3'] = 3
# ws['B1'] = 1
# ws['B2'] = 2
# ws['B3'] = 3

wb.save('sample.xlsx')
