from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(['번호','영어','수학'])
for i in range(1,11): # 데이터 10개 넣기
    ws.append([i,randint(0,100),randint(0,100)])

col_B = ws["B"] # 영어 column만 가져오기
# print(col_B)
for cell in col_B:
    print(cell.value)

col_range = ws['b:c'] #영어, 수학 column 가져오기
print(col_range)

for col in col_range:
    for cell in i:
        print(cell.value)

row_title = ws[1] # 첫번째 row만 가져오기
for row in row_title:
    print(row.value)

# row_range = ws[2:6] # 2줄에서 6번째 줄까지 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end =" ")

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # 2줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.coordinate, end =" ") # 셀의 좌표정보
#         # print(cell.value, end =" ") # 셀값
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy, end = " ")
#         print(xy[0], end = " ") # A
#         print(xy[1], end = " ") # 1
#         # 좌표를 찍어서 작업을 할 때 용이

#     print()

# 전체 rows
# print(tuple(ws.rows))

# 전체 columns
print(tuple(ws.columns))


wb.save('sample.xlsx')