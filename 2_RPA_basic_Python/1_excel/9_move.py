from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active

ws.move_range('b1:c11', rows=0, cols=1)
ws['B1'].value = '국어'

for i in range(2,12):
    ws[f'B{i}'] = randint(50,100)

wb.save('sample_modi.xlsx')