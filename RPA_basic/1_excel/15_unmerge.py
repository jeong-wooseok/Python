from openpyxl import load_workbook
wb = load_workbook('sample_modi.xlsx')
source = wb.active
ws = wb.copy_worksheet(source)
ws.title = 'unmerge'

# B2:D2 병합되어있던 셀을 해제
ws.unmerge_cells('B2:D2')
wb.save('sample_modi.xlsx')