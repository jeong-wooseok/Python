from openpyxl import load_workbook
from openpyxl.drawing.image import Image
wb = load_workbook('sample_modi.xlsx')
ws = wb.create_sheet('image')

img = Image("img.png")

# C3위치에 img 삽입
ws.add_image(img, 'c3')
wb.save('sample_modi.xlsx')