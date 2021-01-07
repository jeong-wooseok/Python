from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, fr
wb = load_workbook('sample.xlsx')
ws = wb.active

# 번호, 영어, 수학
a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']

# 너비/높이 조정
ws.column_dimensions['A'].width = 5 # 열너비 조정
ws.row_dimensions[1].height = 50 # 행높이 설정
wb.save("sample_modi.xlsx")

# 폰트 스타일
a1.font = Font(color='FF0000',italic=True,bold=True) #빨간색, 이탤릭,Bold적용
b1.font = Font(color='CC33FF', name='Arial',strike=True)
c1.font = Font(color='0000FF',size=20,underline='single') #글자크기20, 밑줄적용

# 테두리
thin_border = Border(left=Side(style='thin', color='808080'),\
    right=Side(style='thin', color='808080'),top=Side(style='thin', color='808080'),\
        bottom=Side(style='thin', color='808080'))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 조건부서식 : 90점 넘는 셀에대해 초록색으로 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center') # 가운데 정렬
        if cell.column ==1:
            continue
        # cell이 정수형이고 90점 이상이면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor='00FF00', fill_type='solid')
            cell.font = Font(color='FF0000')

# 틀고정
ws.freeze_panes = 'B2' # B2 기준으로 틀고정

wb.save("sample_modi.xlsx")

